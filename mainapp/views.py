from asyncio.windows_events import NULL
from django.shortcuts import render, redirect
# from mainapp.forms import DataForm
from mainapp.models import HotSales
from mainapp.models import SUM_ACTUAL_PROFIT
from mainapp.models import COST
from mainapp.models import TOTAL_ORDERS
from mainapp.models import MONTH_REVENUE
from mainapp.models import ALL_PROFIT_VW
from mainapp.models import SUM_TOTAL_PROFIT
from mainapp.models import GETYYYYMM
from mainapp.models import GETYYYYMMDD
from mainapp.models import ALLORDERS
from django.shortcuts import render

# Create your views here.
from django import template
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseRedirect
from django.template import loader
from django.urls import reverse
from django.core.files.storage import FileSystemStorage
from django.core.paginator import Paginator,EmptyPage
from django.db import connection
# Create your views here.
from mainapp.serializers import ALLORDERSSerializer
from rest_framework import viewsets




import mysql.connector
import pandas as pd
import xlrd
import csv
import re
import sys
import csv
import requests
import random
import time
import os
import datetime
import json
from datetime import date
from dateutil import relativedelta
from subprocess import run,PIPE
# Create your views here.
class ALLORDERSViewSet(viewsets.ModelViewSet):
    queryset = ALLORDERS.objects.all()
    serializer_class = ALLORDERSSerializer

@login_required(login_url="/login/")
def index(request):
    username = None
    if request.user.is_authenticated:
        username = request.user.username
        if username=='Demo':
            username='florine__20'
            
    ORDERMNMAXSQLstr="select substring(MAX(EFFDT),1,4)*100+substring(MAX(EFFDT),6,2) as YYYYMM from orders Where USERNAME='"+username+"'"
    COSTMNMAXSQLstr="select substring(MAX(YYYYMM),1,4)*100+substring(MAX(YYYYMM),6,2) as YYYYMM from cost Where USERNAME='"+username+"'"
    
    ORDERMNMAX=GETYYYYMM.objects.raw(ORDERMNMAXSQLstr)[0].YYYYMM
    COSTMNMAX=GETYYYYMM.objects.raw(COSTMNMAXSQLstr)[0].YYYYMM
    YYYYMM=min(ORDERMNMAX,COSTMNMAX)
    MM=int(YYYYMM%100)
    YYYY=int((YYYYMM-MM)/100)
    
    Today = date.today() 
    Todayyyymm=str(Today.year)+'-'+str(Today.month)
    abc = datetime.date(YYYY, MM, 1)
    #abc = datetime.date(2022, 10, 31)
    THISYYYYMM=str(abc.year)+'-'+str(abc.month)
    THISYYYY=str(abc.year)
    LASTYYY=str(abc.year-1)
    month, year = (abc.month-1, abc.year) if abc.month != 1 else (12, abc.year-1)
    month2, year2 = (abc.month-2, abc.year) if abc.month != 1 else (12, abc.year-1)
    pre_month = abc.replace(day=1, month=month, year=year)
    pre2_month = abc.replace(day=1, month=month2, year=year2)
    pre_month=str(pre_month)[0:7]
    pre2_month=str(pre2_month)[0:7]
    #GET HOTSALES
    #resultdisplay=HotSales.objects.all()
    HOTSALESSQLstr=  "select D.MAIN_COMMODITY_NM,D.SUMP,D.SUMC,(SUMC-SUMP)/SUMP*100 as PERCENT FROM (select AA.MAIN_COMMODITY_NM,IFNULL(BB.SUMP,0) as SUMP,AA.SUMC FROM (select  A.MAIN_COMMODITY_NM as MAIN_COMMODITY_NM,SUM(A.COUNT) as SUMC  FROM ORDERS A where SUBSTRING(CAST(A.EFFDT AS char),1,7)="
    HOTSALESSQLstr+="'"+pre_month+"'" 
    HOTSALESSQLstr+="and A.STATUS='完成' and USERNAME='"+username+"' group by A.MAIN_COMMODITY_NM order by SUM(A.COUNT) desc LIMIT 5) AA left join (select  B.MAIN_COMMODITY_NM as MAIN_COMMODITY_NM,SUM(B.COUNT) as SUMP  FROM ORDERS B where SUBSTRING(CAST(B.EFFDT AS char),1,7)="
    HOTSALESSQLstr+="'"+pre2_month+"'" 
    HOTSALESSQLstr+="and B.STATUS='完成' and USERNAME='"+username+"' group by B.MAIN_COMMODITY_NM) BB on AA.MAIN_COMMODITY_NM=BB.MAIN_COMMODITY_NM) D"
    resultdisplay=HotSales.objects.raw(HOTSALESSQLstr)
    #GET PROFIT and Margin 
    RevenueSQLstr="select SUM(ACTUAL_PROFIT) as SUM_ACTUAL_PROFIT from ACTUAL_PROFIT where USERNAME='"+username+"' and YYYYMM="
    preRevenueSQLstr=RevenueSQLstr+"'"+pre_month+"'" 
    pre2RevenueSQLstr=RevenueSQLstr+"'"+pre2_month+"'" 
    preRevenue = SUM_ACTUAL_PROFIT.objects.raw(preRevenueSQLstr)[0]
    pre2Revenue = SUM_ACTUAL_PROFIT.objects.raw(pre2RevenueSQLstr)[0]
    CostSQLstr="select USERNAME,YYYYMM,COST from COST where USERNAME='"+username+"' and YYYYMM="
    preCostSQLstr=CostSQLstr+"'"+pre_month+"'" 
    pre2CostSQLstr=CostSQLstr+"'"+pre2_month+"'" 
    preCost = COST.objects.raw(preCostSQLstr)[0]
    pre2Cost = COST.objects.raw(pre2CostSQLstr)[0]
    prePROFIT=int(preRevenue.SUM_ACTUAL_PROFIT-preCost.COST)
    pre2PROFIT=int(pre2Revenue.SUM_ACTUAL_PROFIT-pre2Cost.COST)
    MomProfit=int((prePROFIT-pre2PROFIT)/pre2PROFIT*100)
    
    preMargin=round(prePROFIT/preRevenue.SUM_ACTUAL_PROFIT*100,2)
    pre2Margin=round(pre2PROFIT/pre2Revenue.SUM_ACTUAL_PROFIT*100,2)
    MomMargin=round(preMargin-pre2Margin,2)

    #GET Total Orders
    TotalOrdersSQLstr="select COUNT(1) as COUNT from ACTUAL_PROFIT where USERNAME='"+username+"' and YYYYMM="
    preTotalOrdersSQLstr=TotalOrdersSQLstr+"'"+pre_month+"'" 
    pre2TotalOrdersSQLstr=TotalOrdersSQLstr+"'"+pre2_month+"'" 
    preTotalOrders = TOTAL_ORDERS.objects.raw(preTotalOrdersSQLstr)[0].COUNT
    pre2TotalOrders = TOTAL_ORDERS.objects.raw(pre2TotalOrdersSQLstr)[0].COUNT
    DiffTotalOrders=preTotalOrders-pre2TotalOrders

    #GET REVENUE
    REVENUESQLstr="select YYYYMM,NUMYYYYMM,MONTH_REVENUE from MONTH_REVENUE where USERNAME='"+username+"' and YYYYMM="
    thisREVENUESQLstr=REVENUESQLstr+"'"+THISYYYYMM+"'" 
    thisREVENUE = MONTH_REVENUE.objects.raw(thisREVENUESQLstr)
    if len(list(thisREVENUE))==0:
        thisREVENUEresult=0
    else:
        thisREVENUEresult=int(thisREVENUE[0].MONTH_REVENUE)

    #GET CHART DATA1

 
    numTHISYYYYMM=abc.year*100+abc.month  
    numSixBFYYYYMM=  numTHISYYYYMM-6
    chartREVENUESQLstr="select YYYYMM,NUMYYYYMM,MONTH_REVENUE from MONTH_REVENUE where USERNAME='"+username+"' and NUMYYYYMM>="
    SixBFREVENUESQLstr=chartREVENUESQLstr+str(numSixBFYYYYMM) 
    Chart1Result = MONTH_REVENUE.objects.raw(SixBFREVENUESQLstr)
   
    Chart1Label=[]
    Chart1Data=[]
    Chart2Label=[]
    Month_dict={'01':'JAN','02': 'FEB' ,'03':'MAR' ,'04':'APR' ,'05':'MAY' ,'06':'JUN' ,'07':'JUL','08': 'AUG' ,'09':'SEP','10': 'OCT','11': 'NOV' ,'12':'DEC'}
    for row in Chart1Result:
        YY=row.YYYYMM[2:4]
        MM=row.YYYYMM[5:]
        MN=row.YYYYMM[2:4]+"'"+Month_dict[MM]
        Chart2Label.append(Month_dict[MM])
        Chart1Label.append(MN)
        Chart1Data.append(row.MONTH_REVENUE)

    #GET CHART DATA2

    numLASTYYYYMM=(abc.year-1)*100+abc.month  
    numLASTSixBFYYYYMM=  numLASTYYYYMM-6
    chart2PROFITSQLstr="select YYYYMM,NUMYYYYMM,PROFIT from ALL_PROFIT_VW where USERNAME='"+username+"' and NUMYYYYMM>="
    print(numLASTYYYYMM)
    

    ThisSixBFPROFITSQLstr=chart2PROFITSQLstr+str(numSixBFYYYYMM)+" and NUMYYYYMM<"+str(numTHISYYYYMM)
    LastSixBFPROFITSQLstr=chart2PROFITSQLstr+str(numLASTSixBFYYYYMM) +" and NUMYYYYMM<"+str(numLASTYYYYMM)
    Chart2Result_This = ALL_PROFIT_VW.objects.raw(ThisSixBFPROFITSQLstr)
    Chart2Result_Last = ALL_PROFIT_VW.objects.raw(LastSixBFPROFITSQLstr)
    print(ThisSixBFPROFITSQLstr)

    Chart2Data_This=[]
    Chart2Data_Last=[]

    for row in Chart2Result_This:
        Chart2Data_This.append(row.PROFIT)
    for row in Chart2Result_Last:
        Chart2Data_Last.append(row.PROFIT)

    #GET Profit of The YEAR
    numpre_month=abc.year*100+int(month)
    numlastY_pre_month=(abc.year-1)*100+int(month)
    numpre_month_start=abc.year*100+1
    numlastY_pre_month_start=(abc.year-1)*100+1
    
    ThisTotalPROFITSQLstr="select SUM(PROFIT) as SUM_TOTAL_PROFIT from ALL_PROFIT_VW where USERNAME='"+username+"' and NUMYYYYMM>="
    ThisTotalPROFITSQLstr=ThisTotalPROFITSQLstr+str(numpre_month_start)+" and NUMYYYYMM<="+str(numpre_month)
    TotalPROFIT_This = SUM_TOTAL_PROFIT.objects.raw(ThisTotalPROFITSQLstr)[0]
    TotalPROFIT_This_val=int(TotalPROFIT_This.SUM_TOTAL_PROFIT)

    LastTotalPROFITSQLstr="select SUM(PROFIT) as SUM_TOTAL_PROFIT from ALL_PROFIT_VW where USERNAME='"+username+"' and NUMYYYYMM>="
    LastTotalPROFITSQLstr=LastTotalPROFITSQLstr+str(numlastY_pre_month_start)+" and NUMYYYYMM<="+str(numlastY_pre_month)
    TotalPROFIT_Last = SUM_TOTAL_PROFIT.objects.raw(LastTotalPROFITSQLstr)[0]
    TotalPROFIT_Last_val=int(TotalPROFIT_Last.SUM_TOTAL_PROFIT)
    YOY_TotalPROFIT=round((TotalPROFIT_This_val-TotalPROFIT_Last_val)/TotalPROFIT_Last_val*100,2)


    context = {'segment': 'index','HotSales':resultdisplay,'Title_Pre':pre2_month,'Title_Cur':pre_month,
    'Profit_Cur':prePROFIT,'Profit_Mom':MomProfit,
    'Margin_Cur':preMargin,'Margin_Mom':MomMargin,
    'Total_Orders_Cur':preTotalOrders,'Total_Orders_Dif':DiffTotalOrders,
    'REVENUE_this':thisREVENUEresult,'Chart1Label':Chart1Label,'Chart1Data':Chart1Data,
    'THISYYYY':THISYYYY,'LASTYYY':LASTYYY,
    'Chart2Label':Chart2Label,'Chart2Data_This':Chart2Data_This,'Chart2Data_Last':Chart2Data_Last,
    'TotalPROFIT_This_val':TotalPROFIT_This_val,'YOY_TotalPROFIT':YOY_TotalPROFIT,'Todayyyymm':Todayyyymm
    }
    
    html_template = loader.get_template('home/dashboard.html')
    return HttpResponse(html_template.render(context, request))
    

@login_required(login_url="/login/")
def pages(request):
    context = {}
    # All resource paths end in .html.
    # Pick out the html file name from the url. And load that template.
    try:

        load_template = request.path.split('/')[-1]
        """
        if load_template == 'admin':
            return HttpResponseRedirect(reverse('admin:index'))
        context['segment'] = load_template
        """
        html_template = loader.get_template('home/' + load_template)
        return HttpResponse(html_template.render(context, request))

    except template.TemplateDoesNotExist:

        html_template = loader.get_template('home/page-404.html')
        return HttpResponse(html_template.render(context, request))

    except:
        html_template = loader.get_template('home/page-500.html')
        return HttpResponse(html_template.render(context, request))


@login_required(login_url="/login/")
def orders(request):
    username = None
    demo=False
    if request.user.is_authenticated:
        username = request.user.username
        if username=='Demo':
            username='florine__20'
            demo=True
    ALLORDERSstr=  "SELECT ORDER_ID,EFFDT,STATUS,COMMODITY_NM,SHIIPPING_CODE,SHIPPING_TYPE,RECIPIENT_NAME,PRODUCT_GROSS FROM ALLORDERS Where USERNAME='"+username+"' order by EFFDT desc"
    ALLORDERSDATA=ALLORDERS.objects.raw(ALLORDERSstr)
    for i in range(len(ALLORDERSDATA)):
        ALLORDERSDATA[i].PRODUCT_GROSS=int(ALLORDERSDATA[i].PRODUCT_GROSS)
        ALLORDERSDATA[i].EFFDT=ALLORDERSDATA[i].EFFDT.strftime('%Y-%m-%d %H:%M')
        if demo==True and len(ALLORDERSDATA[i].SHIIPPING_CODE)>0:
            ALLORDERSDATA[i].SHIIPPING_CODE='***' + ALLORDERSDATA[i].SHIIPPING_CODE[3:]
           
    p=Paginator(ALLORDERSDATA,20)
    page_num=request.GET.get('page',1)
    try:
        page=p.page(page_num)
    except EmptyPage:
        page=p.page(1)
    page=p.page(page_num)
    request.session['count'] = 0
    return render(request, 'home/orders.html',{'ALLORDERSDATA':page})

@login_required(login_url="/login/")
def todo(request):
    username = None
    demo=False
    if request.user.is_authenticated:
        username = request.user.username
        if username=='Demo':
            username='florine__20'
            demo=True
    ALLORDERSstr=  "SELECT ORDER_ID,EFFDT,STATUS,COMMODITY_NM,SHIIPPING_CODE,SHIPPING_TYPE,RECIPIENT_NAME,PRODUCT_GROSS FROM ALLORDERS Where STATUS='待出貨' and USERNAME='"+username+"' order by EFFDT desc"
    ALLORDERSDATA=ALLORDERS.objects.raw(ALLORDERSstr)
    for i in range(len(ALLORDERSDATA)):
        ALLORDERSDATA[i].PRODUCT_GROSS=int(ALLORDERSDATA[i].PRODUCT_GROSS)
        ALLORDERSDATA[i].EFFDT=ALLORDERSDATA[i].EFFDT.strftime('%Y-%m-%d %H:%M')
        if demo==True and len(ALLORDERSDATA[i].SHIIPPING_CODE)>0:
            ALLORDERSDATA[i].SHIIPPING_CODE='***' + ALLORDERSDATA[i].SHIIPPING_CODE[3:]
           
    p=Paginator(ALLORDERSDATA,20)
    page_num=request.GET.get('page',1)
    try:
        page=p.page(page_num)
    except EmptyPage:
        page=p.page(1)
    page=p.page(page_num)
    
    return render(request, 'home/todo.html',{'ALLORDERSDATA':page})

@login_required(login_url="/login/")
def simple_upload(request):
    username = None
    if request.user.is_authenticated:
        username = request.user.username
    mydb = mysql.connector.connect(
            host="ec2-54-64-213-252.ap-northeast-1.compute.amazonaws.com",
            user="Eric",
            password="1qaz@WSX",
            database="DEV"
            )
     
    mycursor = mydb.cursor()

    count= request.session.get('count')
    print(count)
    if count==0:
        print('start')
        #GET UploadFromYYYYMMDD
        UPLYYYYMMDDsql="SELECT substring(MIN(EFFDT),1,10) as YYYYMMDD FROM orders where STATUS not in ('完成','不成立')"
        UPLYYYYMMDD=GETYYYYMMDD.objects.raw(UPLYYYYMMDDsql)[0].YYYYMMDD
        
        if UPLYYYYMMDD==None:
            UPLYYYYMMDDsql="SELECT substring(MAX(EFFDT),1,10) as YYYYMMDD FROM orders"
            UPLYYYYMMDD=GETYYYYMMDD.objects.raw(UPLYYYYMMDDsql)[0].YYYYMMDD
            date_object = datetime.datetime.strptime(UPLYYYYMMDD, '%Y-%m-%d').date()
            UPLYYYYMMDD=date_object+datetime.timedelta(days=1)
            UPLYYYYMMDD=UPLYYYYMMDD.strftime('%Y-%m-%d')
        current_dateTime = datetime.datetime.now()
        current_dateTime=current_dateTime.strftime('%Y-%m-%d')
        end_date=datetime.datetime.strptime(current_dateTime, '%Y-%m-%d').date()
        start_date = datetime.datetime.strptime(UPLYYYYMMDD, '%Y-%m-%d').date()
        task=[]
        row=1
        while start_date<end_date:
            s=start_date
            if start_date.month==end_date.month:
                e=end_date
            else:
                e=(datetime.date((start_date+relativedelta.relativedelta(months=1)).year, (start_date+relativedelta.relativedelta(months=1)).month, 1))-datetime.timedelta(days=1)
            sStr=s.strftime('%Y-%m-%d')
            eStr=e.strftime('%Y-%m-%d')
            fs_Str=s.strftime('%Y%m%d')
            fe_Str=e.strftime('%Y%m%d')
            taskStr='Please upload data from '+sStr+' to '+ eStr
            myfilestr='myfile'+str(row)
            uploadstr='upload'+str(row)
            task.append([taskStr,myfilestr,uploadstr,'未完成',fs_Str,fe_Str])
            start_date=e+datetime.timedelta(days=1)
            row=row+1
    
        request.session['task'] = task
        
        deletesql = "delete from ORDERS_UPD_TMP where USERNAME='"+username+"'"
        mycursor.execute(deletesql)
        mydb.commit()
 
    #if request.method == 'POST' and request.FILES['myfile1']:
    if request.method == 'POST':
        task= request.session.get('task')
        for i in range(len(task)):
            mm=task[i][1]
            if request.FILES.get(str(mm))!=None:
                myfile = request.FILES.get(str(mm))
                break
        #myfile = request.FILES['myfile1']
        
        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)
        myfileName=myfile.name
        
        uploaded_file_url = fs.url(filename)
        
        CORE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))   
        XLSX_DIR=os.path.join(CORE_DIR, "media") 
        xlsx_file_name=filename       
        xlsx_file=os.path.join(XLSX_DIR, xlsx_file_name) 
        file_start_date=xlsx_file_name[10:18]
        file_end_date=xlsx_file_name[19:27]
        legal=False
        for i in range(len(task)):
                                      
            if task[i][4]==file_start_date and task[i][5]==file_end_date:
                legal=True
                task[i][3]='完成'
                break
       
       
        rows=[]
        data = xlrd.open_workbook(xlsx_file).sheet_by_index(0)
        
        found=False
        for r in range(data.nrows):
            if r==0:
                continue
            else:     
                temp=[]
                temp.append(username)
                for i in range(0,len(data.row_values(r))):
                    
                    text=data.cell(rowx=r,colx=i).value
                    text=str(text)
                    
                    if (i==43 or i==45 or i==46 or i==47) and len(text)!=16:
                        text=NULL
                    temp.append(text)
                    
        
                rows.append(temp)   
               
           
        
        sql = "INSERT INTO ORDERS_UPD_TMP (USERNAME,ORDER_ID,STATUS,FAIL_REASON,RETURN_STATUS,CUSTOMER_ID,EFFDT,PRODUCT_GROSS,CUSTOMER_SHIP_FEE,SHOPEE_SHIP_FEE,RETURN_SHIP_FEE,TOTAL_PAID,SHOPEE_SUBSIDY,SHOPEE_REDEEM,CREDIT_DISCOUNT,DISCOUNT_CODE,CUSTOMER_COUPON,CUSTOMER_GIVEBACK,SHOPEE_COUPON,TRANSCATION_FEE,SERVICE_FEE,CASHFLOW_FEE,INSTALLMENTS,CREDIT_CHARGE,MAIN_COMMODITY_NM,SUB_COMMODITY_NM,ORG_PRICE,DIS_PRICE,MAIN_COMMODITY_ID,SUB_COMMODITY_ID,COUNT,PROMOTION_METRIC,SHOPEE_PROMOTION,RECIPIENT_ADDR,RECIPIENT_PHONE,PICKUP_STORE_ID,CITY,DISTRICT,POSTAL,RECIPIENT_NAME,SHIPPING_TYPE,DELIVER_TYPE,PREPARE_TIME,PAYMENT_TYPE,LAST_SHIP_TIME,SHIIPPING_CODE,CUSTOMER_PAID_TIME,ACTUAL_SHIP_TIME,ORDER_COMPLETE_TIME,CUSTOMER_COMMENT,COMMENT) VALUES (%s,%s, %s,%s, %s,%s,%s, %s,%s, %s,%s,%s, %s,%s, %s,%s,%s, %s,%s, %s,%s,%s, %s,%s, %s,%s,%s, %s,%s, %s,%s,%s, %s,%s, %s,%s,%s, %s,%s, %s,%s,%s, %s,%s, %s,%s,%s, %s,%s, %s,%s)"
        val = rows
        mycursor.executemany(sql, val)
        mydb.commit()

        mycursor.close()
        #mydb.close()
        os.remove(xlsx_file)
        
        count=count+1
        request.session['count'] = count
        request.session['task'] = task
        print(task)
        print(count)
        return render(request, 'home/simple_upload.html', {
            # 'uploaded_file_url': '/orders',
            # 'uploaded_file_name': myfile.name,
            'count':count,
            'task':task,
            'rows':len(task)
        })
        
    return render(request, 'home/simple_upload.html',{'task':task,'count':count,'rows':len(task)})

@login_required(login_url="/login/")
def update_order(request):
    finish=0
    task= request.session.get('task')
    delete_key_str=str(task[0][4])
    delete_key=delete_key_str[0:4]+'-'+delete_key_str[4:6]+'-'+delete_key_str[6:]
    update_date_str=str(task[-1][5])
    update_date=update_date_str[0:4]+'-'+update_date_str[4:6]+'-'+update_date_str[6:]
    print(delete_key)
    
    username = None
    if request.user.is_authenticated:
        username = request.user.username
    mydb = mysql.connector.connect(
            host="ec2-54-64-213-252.ap-northeast-1.compute.amazonaws.com",
            user="Eric",
            password="1qaz@WSX",
            database="DEV"
            )
     
    mycursor = mydb.cursor()
    deletesql = "delete from ORDERS where USERNAME='"+username+"'"+" and EFFDT>='"+delete_key+"'"
    mycursor.execute(deletesql)
    mydb.commit()
    
    insertsql = "insert into ORDERS (select * from ORDERS_UPD_TMP)"
    mycursor.execute(insertsql)
    mydb.commit()
    mycursor.close()
      
    finish=1
    return render(request, 'home/simple_upload.html',{'redirect':finish,'redirect_url': '/orders','update_date':update_date})

@login_required(login_url="/login/")
def cost(request):
    redirect=0
    username = None
    demo=False
    if request.user.is_authenticated:
        username = request.user.username
        if username=='Demo':
            username='florine__20'
            demo=True
    ALLCOSTSstr=  "SELECT USERNAME,YYYYMM,COST FROM COST Where USERNAME='"+username+"' order by YYYYMM desc"
    ALLCOSTDATA=COST.objects.raw(ALLCOSTSstr)
    ADDLOCK=[]
    row=1
    for i in range(len(ALLCOSTDATA)):
        costid="cost"+str(i+1)
        unlockid="unlock"+str(i+1)
        ADDLOCK.append([ALLCOSTDATA[i].YYYYMM,ALLCOSTDATA[i].COST,costid,unlockid])
        row=row+1
    UPLYYYYMMDD=str(ALLCOSTDATA[0].YYYYMM)+"-01"
    current_dateTime = datetime.datetime.now()
    current_dateTime=current_dateTime.strftime('%Y-%m-%d')
    end_date=datetime.datetime.strptime(current_dateTime, '%Y-%m-%d').date()
    start_date = datetime.datetime.strptime(UPLYYYYMMDD, '%Y-%m-%d').date()
    start_date=(datetime.date((start_date+relativedelta.relativedelta(months=1)).year, (start_date+relativedelta.relativedelta(months=1)).month, 1))
    vals=[]
    while start_date<end_date:
        vals.append([username,start_date.strftime('%Y-%m-%d')[:7],0])
        start_date=(datetime.date((start_date+relativedelta.relativedelta(months=1)).year, (start_date+relativedelta.relativedelta(months=1)).month, 1))
 
    #print(vals)
    for u,y,c in vals:
        costid="cost"+str(row)
        unlockid="unlock"+str(row)
        ADDLOCK.insert(0, [y,c,costid,unlockid])
        row=row+1
    #print(ADDLOCK)
    if len(vals)>0:
        mydb = mysql.connector.connect(
                host="ec2-54-64-213-252.ap-northeast-1.compute.amazonaws.com",
                user="Eric",
                password="1qaz@WSX",
                database="DEV"
                )
        sql = "INSERT INTO COST (USERNAME,YYYYMM,COST) VALUES (%s,%s, %s)"
        mycursor = mydb.cursor()
        mycursor.executemany(sql, vals)
        mydb.commit()
        mycursor.close()
    request.session['ADDLOCK'] = ADDLOCK
    if request.method == 'POST':

        ADDLOCK= request.session.get('ADDLOCK')
        for i in range(len(ADDLOCK)): 
            
            yyyymm=request.POST.get(str(ADDLOCK[i][0]))
            cost=request.POST.get(str(ADDLOCK[i][2]))
            ADDLOCK[i][1]=int(cost)
            print(yyyymm+" "+cost)
            COST.objects.filter(USERNAME=username, YYYYMM=yyyymm).update(COST=cost)
    #     #cost_details=COST(USERNAME=username,YYYYMM= yyyymm,COST = cost)
    #         #defaults = {'COST': 0}
    #         #try:
        
    #         
            #setattr(obj, 'COST', ADDLOCK[i][1])
            # for key, value in defaults.items():
            #     setattr(obj, key, value)
            #obj.save()
            # except COST.DoesNotExist:
            #     print("found")
            #     new_values = {'USERNAME': username, 'YYYYMM': yyyymm,'COST':cost}
            #     new_values.update(defaults)
            #     obj = COST(**new_values)
            #     obj.save()
        """
        cost_details,created=COST.objects.update_or_create(
        USERNAME=username, YYYYMM=yyyymm,COST = cost,
        defaults={'COST': 0},)
        cost_details.save()
        """
        print(ADDLOCK)
        redirect=1
        request.session['ADDLOCK'] = ADDLOCK
        return render(request, 'home/cost.html',
        {'ALLCOSTDATA':ADDLOCK,'redirect':redirect,'redirect_url': '/'}
        )
    
    return render(request, 'home/cost.html',
    {'ALLCOSTDATA':ADDLOCK,'redirect':redirect}
    )

#@login_required(login_url="/login/")
# def button_call(request):
    
#     emoji_pattern = re.compile("["
#                             u"\u2728"
#                             u"\uff5e"
#                             u"\u2763"
#                             u"\ufe0f"
#                             u"\U0001f9f6"
#                             u"\u2661"
#                             u"\U0001f49e"
#                             u"\U0001f60d"
#                             u"\u2764"
#                             u"\U0001f499"
#                             u"\U0001f497"
#                             u"\U0001f970"
#                             u"\U0001f90e"
#                             u"\U0001f90d"
#                             u"\U0001f496"
#                             u"\U0001f929"
#                             u"\U0001f97a"
#                             u"\u9834"
#                             u"\U0001f495"
#                             u"\U0001f493"
#                             u"\U0001f64f"
#                             u"\U0001f3fb"
#                             u"\u51c3"
#                             u"\u5afa"
#                             u"\U0001f49d"
#                             u"\U0001f4e6"
#                             u"\u4f03"
#                             "]+", flags=re.UNICODE)
   

    
#     CORE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))   
#     XLSX_DIR=os.path.join(CORE_DIR, "xlsx") 
#     xlsx_file_name="Order.all.20220701_20220731.xlsx"
#     csv_file_name=xlsx_file_name[:-5]+'.csv'
#     xlsx_file=os.path.join(XLSX_DIR, xlsx_file_name) 
#     csv_file=os.path.join(XLSX_DIR, csv_file_name) 
#     with open(csv_file,'w', newline = "", encoding="big5") as f:
    
#         ob = csv.writer(f)
#         data = xlrd.open_workbook(xlsx_file).sheet_by_index(0)


#         for r in range(data.nrows):
#             if r==0:
#                 HEADER=['ORDER_ID','STATUS','FAIL_REASON','RETURN_STATUS','CUSTOMER_ID','EFFDT','PRODUCT_GROSS','CUSTOMER_SHIP_FEE','SHOPEE_SHIP_FEE','RETURN_SHIP_FEE','TOTAL_PAID','SHOPEE_SUBSIDY','SHOPEE_REDEEM','CREDIT_DISCOUNT','DISCOUNT_CODE','CUSTOMER_COUPON','CUSTOMER_GIVEBACK','SHOPEE_COUPON','TRANSCATION_FEE','SERVICE_FEE','CASHFLOW_FEE','INSTALLMENTS','CREDIT_CHARGE','MAIN_COMMODITY_NM','SUB_COMMODITY_NM','ORG_PRICE','DIS_PRICE','MAIN_COMMODITY_ID','SUB_COMMODITY_ID','COUNT','PROMOTION_METRIC','SHOPEE_PROMOTION','RECIPIENT_ADDR','RECIPIENT_PHONE','PICKUP_STORE_ID','CITY','DISTRICT','POSTAL','RECIPIENT_NAME','SHIPPING_TYPE','DELIVER_TYPE','PREPARE_TIME','PAYMENT_TYPE','LAST_SHIP_TIME','SHIIPPING_CODE','CUSTOMER_PAID_TIME','ACTUAL_SHIP_TIME','ORDER_COMPLETE_TIME','CUSTOMER_COMMENT','COMMENT']
#                 ob.writerow(HEADER)
#             else:     
#                 temp=[]
#                 for i in range(0,len(data.row_values(r))):

#                     text=data.cell(rowx=r,colx=i).value
#                     text=str(text)
#                     text= emoji_pattern.sub(r'', text)
#                     temp.append(text)
#                 ob.writerow(temp)   
    
#     #print("Hello World!")
#     #context = {'segment': 'index'}

#     #html_template = loader.get_template('home/dashboard.html')
#     #return HttpResponse(html_template.render(context, request))
#     #out=run(sys.executable,['D:\\Eric\\AWS\\AWS_FLORINE\\AWS_FLORINE_DJ\\apps\home\\function.py'],shell=False,stdout=PIPE)
#     #out="HI"
#     #print(out)

#     return render(request,'home/dashboard.html')

"""
def index(request):
    form = DataForm()
    queryset = models.Data.objects.all()
    if request.method == 'POST':
        form = DataForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('index')
    context = {'title': 'Simple App', 'form': form, 'posts': queryset}
    return render(request, 'mainapp/index.html', context=context)


from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.styles import Font

def export_data(request):
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Data.xlsx"'

    # create workbook
    wb = Workbook()
    sheet = wb.active

    # stylize header row
    # 'id','title', 'quantity','pub_date'
      
    c1 = sheet.cell(row = 1, column = 1) 
    c1.value = "id"
    c1.font = Font(bold=True)
    
    c2 = sheet.cell(row= 1 , column = 2) 
    c2.value = "title"
    c2.font = Font(bold=True)

    c3 = sheet.cell(row= 1 , column = 3) 
    c3.value = "quantity"
    c3.font = Font(bold=True)

    c4 = sheet.cell(row= 1 , column = 4) 
    c4.value = "pub_date"
    c4.font = Font(bold=True)
    
    # export data to Excel
    rows = models.Data.objects.all().values_list('id','category', 'quantity','pub_date',)
    for row_num, row in enumerate(rows, 1):
        # row is just a tuple
        for col_num, value in enumerate(row):
            c5 = sheet.cell(row=row_num+1, column=col_num+1) 
            c5.value = value

    wb.save(response)

    return response
"""